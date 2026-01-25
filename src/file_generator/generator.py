"""
File Generator Module
Generates bait files with embedded tracking beacons.
"""
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import uuid
from datetime import datetime
from faker import Faker
import logging

fake = Faker()


class FileGenerator:
    """Generate tracked bait files (PDF, Excel, config files)"""
    
    def __init__(self, server_url="http://localhost:8000", output_dir="generated_files"):
        """
        Initialize file generator
        
        Args:
            server_url: Your honeypot URL for tracking callbacks
            output_dir: Directory to save generated files
        """
        self.server_url = server_url
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def generate_pdf(self, filename, client_ip, content_data=None):
        """
        Generate PDF with embedded tracking beacon
        
        Args:
            filename: Original filename requested
            client_ip: IP of requester
            content_data: Optional dict with content from LLM
            
        Returns:
            Path to generated file and beacon_id
        """
        beacon_id = str(uuid.uuid4())
        filepath = self.output_dir / f"{beacon_id}.pdf"
        
        try:
            c = canvas.Canvas(str(filepath), pagesize=letter)
            width, height = letter
            
            # Header
            c.setFont("Helvetica-Bold", 18)
            c.drawString(1*inch, height - 1*inch, "CONFIDENTIAL - Internal Document")
            
            # Metadata
            c.setFont("Helvetica", 10)
            c.drawString(1*inch, height - 1.3*inch, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            c.drawString(1*inch, height - 1.5*inch, f"Document ID: {beacon_id[:8]}")
            c.drawString(1*inch, height - 1.7*inch, f"Classification: INTERNAL USE ONLY")
            
            # Content
            c.setFont("Helvetica-Bold", 14)
            c.drawString(1*inch, height - 2.2*inch, "Q4 Financial Report - 2024")
            
            c.setFont("Helvetica", 11)
            y_position = height - 2.6*inch
            
            if content_data:
                # Use LLM-generated content
                for key, value in content_data.items():
                    c.drawString(1*inch, y_position, f"{key}: {value}")
                    y_position -= 0.3*inch
            else:
                # Default content
                lines = [
                    "Revenue: $4,234,567",
                    "Operating Expenses: $2,123,456",
                    "Net Profit: $2,111,111",
                    "Growth Rate: 23.4%",
                    "",
                    "Key Metrics:",
                    "  - Customer Acquisition Cost: $234",
                    "  - Lifetime Value: $1,234",
                    "  - Monthly Recurring Revenue: $456,789",
                    "",
                    "Prepared by: Finance Department",
                    f"Contact: finance@corporate.internal"
                ]
                
                for line in lines:
                    c.drawString(1*inch, y_position, line)
                    y_position -= 0.25*inch
            
            # TRACKING BEACON - Invisible image that loads from honeypot
            tracking_url = f"{self.server_url}/track/{beacon_id}"
            
            # Note: ReportLab's drawImage needs a local file for external URLs
            # Alternative: Embed beacon ID in metadata
            c.setAuthor(f"Beacon-{beacon_id}")
            c.setTitle(filename)
            c.setSubject(f"Tracking: {tracking_url}")
            
            # Add visible tracking (looks like a logo URL)
            c.setFont("Helvetica", 6)
            c.setFillColorRGB(0.9, 0.9, 0.9)  # Very light gray
            c.drawString(1*inch, 0.5*inch, f"Document verification: {tracking_url}")
            
            c.save()
            
            self.logger.info(f"Generated PDF: {filename} with beacon {beacon_id} for {client_ip}")
            return filepath, beacon_id
            
        except Exception as e:
            self.logger.error(f"PDF generation failed: {e}")
            raise
    
    def generate_excel(self, filename, client_ip, content_data=None):
        """
        Generate Excel file with tracking in metadata
        
        Args:
            filename: Original filename
            client_ip: IP of requester
            content_data: Optional dict/list from LLM
            
        Returns:
            Path to generated file and beacon_id
        """
        beacon_id = str(uuid.uuid4())
        filepath = self.output_dir / f"{beacon_id}.xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Employee Directory"
            
            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Headers
            headers = ["ID", "Name", "Email", "Department", "Salary", "Hire Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Data
            if content_data and isinstance(content_data, list):
                # Use LLM-generated data
                for row_idx, employee in enumerate(content_data, 2):
                    ws.cell(row=row_idx, column=1, value=employee.get('id', row_idx-1))
                    ws.cell(row=row_idx, column=2, value=employee.get('name', fake.name()))
                    ws.cell(row=row_idx, column=3, value=employee.get('email', fake.email()))
                    ws.cell(row=row_idx, column=4, value=employee.get('department', fake.job()))
                    ws.cell(row=row_idx, column=5, value=employee.get('salary', fake.random_int(50000, 150000)))
                    ws.cell(row=row_idx, column=6, value=employee.get('hire_date', fake.date_this_decade()))
            else:
                # Generate fake data
                for i in range(2, 22):  # 20 employees
                    ws.cell(row=i, column=1, value=i-1)
                    ws.cell(row=i, column=2, value=fake.name())
                    ws.cell(row=i, column=3, value=fake.email())
                    ws.cell(row=i, column=4, value=fake.job())
                    ws.cell(row=i, column=5, value=fake.random_int(50000, 150000))
                    ws.cell(row=i, column=6, value=fake.date_this_decade())
            
            # Adjust column widths
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            
            # TRACKING BEACON - Embedded in metadata
            tracking_url = f"{self.server_url}/track/{beacon_id}"
            wb.properties.creator = f"Beacon-{beacon_id}"
            wb.properties.company = "Corporate Internal"
            wb.properties.comments = f"Verification URL: {tracking_url}"
            wb.properties.title = filename
            
            wb.save(filepath)
            
            self.logger.info(f"Generated Excel: {filename} with beacon {beacon_id} for {client_ip}")
            return filepath, beacon_id
            
        except Exception as e:
            self.logger.error(f"Excel generation failed: {e}")
            raise
    
    def generate_env_file(self, filename, client_ip, content_data=None):
        """
        Generate .env configuration file with fake credentials
        
        Args:
            filename: Original filename
            client_ip: IP of requester
            content_data: Optional dict from LLM
            
        Returns:
            Path to generated file and beacon_id
        """
        beacon_id = str(uuid.uuid4())
        filepath = self.output_dir / f"{beacon_id}.env"
        
        try:
            tracking_url = f"{self.server_url}/track/{beacon_id}"
            
            if content_data and isinstance(content_data, dict):
                # Use LLM-generated content
                lines = [f"# Production Environment Configuration"]
                lines.append(f"# Generated: {datetime.now().isoformat()}")
                lines.append(f"# Beacon: {beacon_id}")
                lines.append("")
                
                for key, value in content_data.items():
                    lines.append(f"{key}={value}")
            else:
                # Default fake credentials
                lines = [
                    f"# Production Environment Configuration",
                    f"# Generated: {datetime.now().isoformat()}",
                    f"# Document ID: {beacon_id}",
                    "",
                    "# Database",
                    "DATABASE_URL=postgresql://admin:P@ssw0rd123@db.internal.corp:5432/production",
                    "DB_HOST=10.0.1.50",
                    "DB_USER=dbadmin",
                    "DB_PASSWORD=SuperSecret123!",
                    "",
                    "# Redis Cache",
                    "REDIS_URL=redis://cache.internal.corp:6379",
                    "REDIS_PASSWORD=redis_pass_456",
                    "",
                    "# API Keys",
                    "API_KEY=sk-prod-FAKE1234567890ABCDEFGHIJKLMNOP",
                    "SECRET_KEY=super_secret_key_do_not_share_789xyz",
                    "JWT_SECRET=jwt_secret_token_456def",
                    "",
                    "# AWS Credentials",
                    "AWS_ACCESS_KEY_ID=AKIA1234FAKE5678KEYS",
                    "AWS_SECRET_ACCESS_KEY=wJalrXUtnFEMI/K7MDENG/bPxRfiCYFAKEKEY",
                    "AWS_REGION=us-east-1",
                    "S3_BUCKET=corporate-production-data",
                    "",
                    "# External Services",
                    "STRIPE_KEY=sk_live_FAKE_stripe_key_123456",
                    "SENDGRID_API_KEY=SG.FAKE_sendgrid_key_789",
                    "",
                    f"# Tracking Beacon: {tracking_url}",
                ]
            
            content = "\n".join(lines)
            
            with open(filepath, 'w') as f:
                f.write(content)
            
            self.logger.info(f"Generated .env: {filename} with beacon {beacon_id} for {client_ip}")
            return filepath, beacon_id
            
        except Exception as e:
            self.logger.error(f".env generation failed: {e}")
            raise
    
    def generate_config_file(self, filename, client_ip, file_format="yaml"):
        """
        Generate configuration file (YAML, JSON, INI)
        
        Args:
            filename: Original filename
            client_ip: IP of requester
            file_format: yaml, json, or ini
            
        Returns:
            Path to generated file and beacon_id
        """
        beacon_id = str(uuid.uuid4())
        filepath = self.output_dir / f"{beacon_id}.{file_format}"
        
        tracking_url = f"{self.server_url}/track/{beacon_id}"
        
        if file_format == "yaml":
            content = f"""# Production Configuration
# Beacon: {beacon_id}
server:
  host: 0.0.0.0
  port: 8080
  workers: 4
database:
  host: db.internal.corp
  port: 5432
  username: admin
  password: DbPass123!
  database: production
cache:
  type: redis
  host: cache.internal.corp
  password: RedisPass456
security:
  secret_key: "super-secret-key-xyz789"
  jwt_expiry: 3600
  
# Tracking: {tracking_url}
"""
        elif file_format == "json":
            content = f'''{{
  "_comment": "Production Configuration - Beacon: {beacon_id}",
  "server": {{
    "host": "0.0.0.0",
    "port": 8080,
    "workers": 4
  }},
  "database": {{
    "host": "db.internal.corp",
    "username": "admin",
    "password": "DbPass123!",
    "database": "production"
  }},
  "tracking_url": "{tracking_url}"
}}'''
        else:  # ini
            content = f"""[server]
host = 0.0.0.0
port = 8080
[database]
host = db.internal.corp
username = admin
password = DbPass123!
[tracking]
url = {tracking_url}
beacon = {beacon_id}
"""
        
        with open(filepath, 'w') as f:
            f.write(content)
        
        self.logger.info(f"Generated {file_format}: {filename} with beacon {beacon_id} for {client_ip}")
        return filepath, beacon_id
    
    def generate_random_file(self, endpoint_path, client_ip, rag_context=None):
        """
        Generate a random file type based on endpoint context
        
        Args:
            endpoint_path: The API endpoint path requesting the file
            client_ip: IP of requester
            rag_context: Optional RAG context for realistic content
            
        Returns:
            tuple: (filepath, filename, beacon_id, file_type)
        """
        import random
        from .sqlite_gen import SQLiteGenerator
        from .txt_gen import TextFileGenerator
        
        beacon_id = str(uuid.uuid4())
        endpoint_lower = endpoint_path.lower()
        
        # Determine file type based on endpoint
        if 'database' in endpoint_lower or 'db' in endpoint_lower or 'data' in endpoint_lower:
            # Prefer SQLite for database endpoints
            file_types = ['sqlite', 'excel', 'pdf']
            weights = [0.6, 0.2, 0.2]
        elif 'config' in endpoint_lower or 'settings' in endpoint_lower:
            # Prefer config files
            file_types = ['txt', 'env', 'yaml']
            weights = [0.4, 0.4, 0.2]
        elif 'secret' in endpoint_lower or 'cred' in endpoint_lower or 'key' in endpoint_lower:
            # Prefer credentials
            file_types = ['txt', 'env']
            weights = [0.6, 0.4]
        elif 'log' in endpoint_lower or 'audit' in endpoint_lower:
            # Prefer logs
            file_types = ['txt', 'pdf']
            weights = [0.7, 0.3]
        else:
            # Random mix
            file_types = ['pdf', 'txt', 'sqlite', 'env', 'excel']
            weights = [0.3, 0.25, 0.2, 0.15, 0.1]
        
        file_type = random.choices(file_types, weights=weights)[0]
        
        try:
            if file_type == 'sqlite':
                sql_gen = SQLiteGenerator(output_dir=str(self.output_dir / "databases"))
                filepath, filename = sql_gen.generate_database(rag_context, beacon_id, endpoint_path)
            elif file_type == 'txt':
                txt_gen = TextFileGenerator(output_dir=str(self.output_dir / "textfiles"))
                filepath, filename = txt_gen.generate_text_file(rag_context, beacon_id, endpoint_path)
            elif file_type == 'pdf':
                filepath, _ = self.generate_pdf(f"document_{beacon_id[:8]}.pdf", client_ip, rag_context)
                filename = filepath.name
            elif file_type == 'env':
                filepath, _ = self.generate_env_file(f"config_{beacon_id[:8]}.env", client_ip, rag_context)
                filename = filepath.name
            elif file_type == 'excel':
                filepath, _ = self.generate_excel(f"data_{beacon_id[:8]}.xlsx", client_ip, rag_context)
                filename = filepath.name
            elif file_type == 'yaml':
                filepath, _ = self.generate_config_file(f"config_{beacon_id[:8]}.yaml", client_ip, "yaml")
                filename = filepath.name
            else:
                # Fallback to PDF
                filepath, _ = self.generate_pdf(f"document_{beacon_id[:8]}.pdf", client_ip, rag_context)
                filename = filepath.name
            
            self.logger.info(f"Generated random {file_type} file: {filename} for endpoint {endpoint_path}")
            return filepath, filename, beacon_id, file_type
            
        except Exception as e:
            self.logger.error(f"Random file generation failed: {e}")
            # Fallback to simple PDF
            filepath, _ = self.generate_pdf(f"fallback_{beacon_id[:8]}.pdf", client_ip)
            return filepath, filepath.name, beacon_id, 'pdf'
